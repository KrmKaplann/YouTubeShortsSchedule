import json
import os
import undetected_chromedriver as uc
import time
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import pyautogui
import locale
from datetime import datetime

locale.setlocale(locale.LC_TIME, "tr_TR.UTF-8")

def Driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    #options.add_argument("--incognito")
    #options.add_argument("--headless")  # Run the browser in headless mode
    options.add_argument("--start-maximized")  # Start the browser maximized
    driver = uc.Chrome(options=options)
    return driver

def Chrome_Tab(driver, link):
    driver.get(link)
    driver.maximize_window()

path = "<ExcelFilePath>"
# <ExcelFilePath>: Path to the Excel file
workbook = openpyxl.load_workbook(path)
DefaultPageSheet = workbook["<SheetName>"]
# <SheetName>: Name of the sheet in the Excel file

IslemYapilacakSosyalMedya = "YouTube"

StartColumn = 5
while True:
    Finder = DefaultPageSheet.cell(2, StartColumn).value
    if Finder == IslemYapilacakSosyalMedya:
        break
    StartColumn += 1

TotalAccountListGmail = []
DefaultRow = 4

while True:
    Start = DefaultPageSheet.cell(DefaultRow, StartColumn).value
    if Start is None:
        break
    if DefaultPageSheet.cell(DefaultRow, StartColumn + 1).value != "-":
        TotalAccountListGmail.append([Start, DefaultRow - 3])
    DefaultRow += 1

# Filter out entries with '-'
TotalAccountListGmail = [item for item in TotalAccountListGmail if item[0] != '-']

TotalAccountList = []
DefaultRow = 4

for account_info in TotalAccountListGmail:
    index = account_info[1]
    Start = DefaultPageSheet.cell(index + 3, 2).value
    TotalAccountList.append(Start)

print(TotalAccountList)

# Get the starting index from the user
baslangic_indeksi = int(input("Enter the last completed index for YouTube accounts (0 to {}): ".format(len(TotalAccountList)-1)))

TotalAccountList = TotalAccountList[baslangic_indeksi:]
TotalAccountListGmail = TotalAccountListGmail[baslangic_indeksi:]

path = "<YouTubeExcelFilePath>"
# <YouTubeExcelFilePath>: Path to the YouTube Excel file
workbook = openpyxl.load_workbook(path)
AllPagesWorksheet = workbook.sheetnames

for IndexAccount, OneItem in enumerate(TotalAccountListGmail, start=0):
    driver = Driver()
    link = "https://www.youtube.com"

    Chrome_Tab(driver, link)

    JsonFileName = DefaultPageSheet.cell(2, StartColumn).value + OneItem[0] + ".json"
    with open(f"<APIsPath>/{JsonFileName}", "r") as file:
        # <APIsPath>: Path to the APIs directory
        cookies = json.load(file)

    for cookie in cookies:
        driver.add_cookie(cookie)

    time.sleep(3)
    driver.refresh()

    for index, OnePage in enumerate(TotalAccountList, start=1):
        print(f"[ {index} ] {OnePage}")

    StartRow = 5
    PostDict = {}

    sekmeler = driver.window_handles
    ilk_sekme = sekmeler[0]
    driver.switch_to.window(ilk_sekme)

    time.sleep(2)
    if StartRow == 5:
        for i in range(3):
            time.sleep(1)
            pyautogui.hotkey('ctrl', '-')
    time.sleep(1)

    OneAccountName = TotalAccountList[IndexAccount]
    StartRow = 5
    AllPostList = []
    while True:
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        StartRow += 1
        if PostName is None:
            break
        AllPostList.append(PostName)

    print(AllPostList)
    AllPostDefault = AllPostList
    baslangic_indeksi = int(input(f"Enter the last completed index for {OneAccountName} (0 to {(len(AllPostList) - 1)}): "))

    AllPostList = AllPostList[baslangic_indeksi:]

    for StartRow, PostOne in enumerate(AllPostDefault, start=5):
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        Aciklama = SpecialWorkSheet["C" + str(StartRow)].value

        Tarih = SpecialWorkSheet["D" + str(StartRow)].value
        Tarih = datetime.strptime(Tarih, '%d.%m.%Y')
        FormatliTarih = Tarih.strftime("%d %B %A %Y")
        Gun = Tarih.strftime("%d")
        Ay = Tarih.strftime("%B")
        Saat = str(SpecialWorkSheet["E" + str(StartRow)].value)
        SaatSplitted = Saat.split(":")
        Hour = str(SaatSplitted[0])
        Minutes = str(SaatSplitted[1])
        HemenPaylas = SpecialWorkSheet["F" + str(StartRow)].value

        post_details = {
            "Aciklama": Aciklama,
            "Tarih": FormatliTarih,
            "Gün": Gun,
            "Ay": Ay,
            "Saat": Saat,
            "Hour": Hour,
            "Minutes": Minutes,
            "HemenPaylas": HemenPaylas
        }

        if OneAccountName in PostDict:
            PostDict[OneAccountName][PostName] = post_details
        else:
            PostDict[OneAccountName] = {PostName: post_details}

    for StartRow, PostOne in enumerate(AllPostList, start=5):
        wait = WebDriverWait(driver, 60)
        driver.implicitly_wait(10)
        LogoVisibility = wait.until(EC.visibility_of_element_located((By.XPATH, "<AvatarButton>")))
        # <AvatarButton>: XPath for avatar button
        time.sleep(2)
        driver.find_element(By.XPATH, "<CreateButton>").click()
        # <CreateButton>: XPath for create button
        time.sleep(2)
        driver.find_element(By.XPATH, "<UploadVideoButton>").click()
        # <UploadVideoButton>: XPath for upload video button
        time.sleep(2)
        try:
            driver.find_element(By.XPATH, "<CloseButton>").click()
            # <CloseButton>: XPath for close button
        except:
            pass
        time.sleep(1)
        driver.find_element(By.XPATH, "<UploadButton>").click()
        # <UploadButton>: XPath for upload button
        time.sleep(2)
        driver.find_element(By.XPATH, "<BetaUploadButton>").click()
        # <BetaUploadButton>: XPath for beta upload button
        time.sleep(2)
        FileAdress = "<YouTubeVideoPath>/" + OneAccountName + "/" + PostOne + ".mp4"
        # <YouTubeVideoPath>: Path to the YouTube videos
        time.sleep(3)
        driver.find_element(By.XPATH, "<FileInput>").send_keys(FileAdress)
        # <FileInput>: XPath for file input

        TextAreaVisibility = wait.until(EC.visibility_of_element_located((By.XPATH, "<TitleTextarea>")))
        # <TitleTextarea>: XPath for title textarea

        Aciklama = str(PostDict[OneAccountName][PostOne]["Aciklama"])
        if "/Başlık" in Aciklama:
            Baslik = (Aciklama.split("/Başlık"))[0]
            UzunAciklama = (Aciklama.split("/Başlık"))[1]
            time.sleep(2)
            driver.find_element(By.XPATH, "<TitleTextarea>").send_keys(Keys.CONTROL+"a")
            driver.find_element(By.XPATH, "<TitleTextarea>").send_keys(Baslik)
            driver.find_element(By.XPATH, "<DescriptionTextarea>").send_keys(UzunAciklama)
            # <DescriptionTextarea>: XPath for description textarea
        else:
            driver.find_element(By.XPATH, "<TitleTextarea>").send_keys(Keys.CONTROL+"a")
            driver.find_element(By.XPATH, "<TitleTextarea>").send_keys(str(PostDict[OneAccountName][PostOne]["Aciklama"]))

        driver.find_element(By.XPATH, "<VideoMadeForKidsButton>").click()
        # <VideoMadeForKidsButton>: XPath for video made for kids button
        time.sleep(3)
        for i in range(3):
            driver.find_element(By.XPATH, "<NextButton>").click()
            # <NextButton>: XPath for next button
            time.sleep(2)

        time.sleep(2)
        driver.find_element(By.XPATH, "<ExpandButton>").click()
        # <ExpandButton>: XPath for expand button
        time.sleep(2)
        driver.find_element(By.XPATH, "<DatepickerTrigger>").click()
        # <DatepickerTrigger>: XPath for datepicker trigger
        time.sleep(2)
        CombinedDate = datetime.strptime(str(PostDict[OneAccountName][PostOne]["Tarih"]), '%d %B %A %Y').strftime('%d.%m.%Y')

        driver.find_element(By.XPATH, "<DateInput>").send_keys(Keys.CONTROL + "a")
        # <DateInput>: XPath for date input
        time.sleep(2)
        driver.find_element(By.XPATH, "<DateInput>").send_keys(CombinedDate)
        driver.find_element(By.XPATH, "<DateInput>").send_keys(Keys.ENTER)

        time.sleep(2)
        SaatVEDakika = str(PostDict[OneAccountName][PostOne]["Hour"]) + ":" + str(PostDict[OneAccountName][PostOne]["Minutes"])
        driver.find_element(By.XPATH, "<TimeInput>").send_keys(Keys.CONTROL+"a")
        # <TimeInput>: XPath for time input
        driver.find_element(By.XPATH, "<TimeInput>").send_keys(SaatVEDakika)
        time.sleep(1)

        time.sleep(10)
        driver.find_element(By.XPATH, "<ScheduleButton>").click()
        # <ScheduleButton>: XPath for schedule button
        time.sleep(10)
        print(f"{PostOne} completed...")
        driver.get("https://www.youtube.com")
        time.sleep(3)

    driver.quit()

print("Completed...")
